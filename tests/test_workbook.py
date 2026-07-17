"""The Excel workbook builds from the synthetic sample data and contains the
expected category/master tabs."""

from pathlib import Path

from openpyxl import load_workbook

from ig_inbox import build_workbook, build_lists

EXAMPLES = Path(__file__).resolve().parent.parent / "examples"
SAMPLE = EXAMPLES / "captures.jsonl"
BOOKS = EXAMPLES / "books.master.json"


def test_workbook_builds_from_sample(tmp_path):
    out = tmp_path / "workbook.xlsx"
    n = build_workbook.build(captures_path=SAMPLE, out_path=out, books_master=BOOKS,
                             known_games={"palworld": "Palworld", "pow world": "Palworld"})
    assert n == 10
    assert out.exists()

    wb = load_workbook(out)
    names = set(wb.sheetnames)
    # Overview + masters + a spread of category/collection tabs.
    assert "Overview" in names
    assert "Books" in names          # deduped master
    assert "Restaurants" in names    # deduped master (from the collection)
    assert "Recipes" in names
    assert "AI Ideas" in names
    assert "Finance" in names

    # Books master carries the enriched rows.
    ws = wb["Books"]
    header = [c.value for c in ws[1]]
    assert "Title" in header and "Rating" in header
    titles = [ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)]
    assert "The Long Horizon" in titles


def test_workbook_handles_empty_captures(tmp_path):
    empty = tmp_path / "empty.jsonl"
    empty.write_text("")
    out = tmp_path / "wb.xlsx"
    n = build_workbook.build(captures_path=empty, out_path=out,
                             books_master=tmp_path / "nope.json", known_games={})
    assert n == 0
    assert out.exists()  # still writes a valid (Overview-only) workbook


def test_lists_build_from_sample(tmp_path):
    lists_dir = tmp_path / "lists"
    n = build_lists.build_all(captures_path=SAMPLE, lists_dir=lists_dir)
    assert n == 10
    assert (lists_dir / "Books.md").exists()
    assert (lists_dir / "AI Ideas.md").exists()
    assert (lists_dir / "Finance.md").exists()
    # The finance list should surface the tickers from the sample.
    finance = (lists_dir / "Finance.md").read_text()
    assert "NVDA" in finance
