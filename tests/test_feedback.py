"""The correction feedback loop: capture → log → pin → few-shot → metrics."""

import json
from pathlib import Path

from openpyxl import load_workbook

from ig_inbox import build_workbook, config, feedback, route_store

EXAMPLES = Path(__file__).resolve().parent.parent / "examples"
SAMPLE = EXAMPLES / "captures.jsonl"


def _seed(tmp_path):
    """Copy the sample captures into the test data dir so feedback can mutate it.
    Also reset feedback state so tests don't leak into each other (the data dir
    is shared for the whole session)."""
    config.DATA_DIR.mkdir(parents=True, exist_ok=True)
    for f in (config.CORRECTIONS_FILE, config.CORRECTIONS_INBOX, config.METRICS_FILE):
        f.unlink(missing_ok=True)
    recs = [json.loads(l) for l in SAMPLE.read_text().split("\n") if l.strip()]
    route_store.rewrite_captures(recs, captures=config.CAPTURES_FILE)
    return recs


def test_resolve_category_matches_keys_labels_and_coins(monkeypatch, tmp_path):
    monkeypatch.setattr(config, "TAXONOMY_FILE", tmp_path / "tax.json")
    assert config.resolve_category("recipe") == "recipe"        # seed key
    assert config.resolve_category("Recipes") == "recipe"       # pretty label → key
    assert config.resolve_category("Woodworking") == "woodworking"  # coins a new one
    assert config.resolve_category("") is None                  # empty → nothing


def test_inbox_correction_logs_pins_and_learns(tmp_path):
    _seed(tmp_path)
    # The sample "standing-desk converter" note is category 'product'. Say the
    # user re-files it as 'career' via the manual inbox path.
    config.CORRECTIONS_INBOX.write_text(
        json.dumps({"id": "ig_sample_0010", "to_category": "career"}) + "\n")

    n = feedback.ingest_corrections()
    assert n == 1

    # 1) correction is logged durably
    logged = [json.loads(l) for l in config.CORRECTIONS_FILE.read_text().split("\n") if l.strip()]
    assert logged[-1]["from_category"] == "product"
    assert logged[-1]["to_category"] == "career"

    # 2) the record is pinned to the new category in captures.jsonl
    recs = {r["id"]: r for r in route_store._read_records(config.CAPTURES_FILE)}
    assert recs["ig_sample_0010"]["category"] == "career"
    assert recs["ig_sample_0010"]["pinned"] is True

    # 3) it becomes a few-shot hint for the classifier
    block = feedback.few_shot_block()
    assert "Career" in block and "should be" in block

    # inbox file is consumed
    assert not config.CORRECTIONS_INBOX.exists()


def test_metrics_populate(tmp_path):
    _seed(tmp_path)
    config.CORRECTIONS_INBOX.write_text(
        json.dumps({"id": "ig_sample_0010", "to_category": "career"}) + "\n")
    feedback.ingest_corrections()

    m = feedback.compute_metrics()
    assert m["total_items_classified"] == 10
    assert m["total_corrections"] == 1
    assert m["correction_rate"] == round(1 / 10, 3)
    assert "Products → Career" in m["by_transition"]
    assert config.METRICS_FILE.exists()


def test_workbook_has_metrics_tab_and_correction_columns(tmp_path):
    _seed(tmp_path)
    out = tmp_path / "wb.xlsx"
    build_workbook.build(captures_path=config.CAPTURES_FILE, out_path=out,
                         books_master=EXAMPLES / "books.master.json", known_games={})
    wb = load_workbook(out)
    assert "Metrics" in wb.sheetnames
    # a per-post sheet must expose the editable correction + id columns
    other = wb["Products"] if "Products" in wb.sheetnames else wb["Other"]
    header = [c.value for c in other[1]]
    assert feedback.CORRECTION_HEADER in header
    assert feedback.ID_HEADER in header


def test_workbook_override_readback(tmp_path):
    """Simulate a user typing a category into the workbook, then reading it back."""
    _seed(tmp_path)
    out = config.WORKBOOK_FILE
    build_workbook.build(captures_path=config.CAPTURES_FILE, out_path=out,
                         books_master=EXAMPLES / "books.master.json", known_games={})
    # user edits the workbook: put "Career" on the standing-desk row
    wb = load_workbook(out)
    ws = wb["Products"]
    header = [c.value for c in ws[1]]
    ci_corr = header.index(feedback.CORRECTION_HEADER) + 1
    ci_id = header.index(feedback.ID_HEADER) + 1
    for r in range(2, ws.max_row + 1):
        if ws.cell(row=r, column=ci_id).value == "ig_sample_0010":
            ws.cell(row=r, column=ci_corr, value="Career")
    wb.save(out)

    overrides = feedback.read_workbook_overrides(out)
    assert overrides.get("ig_sample_0010") == "career"
