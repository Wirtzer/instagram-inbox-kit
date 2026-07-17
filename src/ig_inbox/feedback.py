"""Self-improving feedback loop + telemetry.

When the user re-categorizes an item — by typing a category into the workbook's
"Correct category?" column, or by dropping a line into data/corrections.inbox.jsonl
— this module:

  1. Reads those overrides BEFORE the workbook is regenerated (the workbook is
     rebuilt every run, so overrides MUST be harvested first or they'd be lost).
  2. Records each real change to data/corrections.jsonl (append-only, atomic,
     locked — same durability as captures).
  3. Pins the corrected record to the user's chosen category (never re-flipped)
     in captures.jsonl.
  4. Feeds the accumulated corrections back into the classifier as compact
     few-shot examples (see few_shot_block), so similar future items land right.
  5. Recomputes data/metrics.json (totals, transitions, correction rate over
     time, most-corrected categories, a "what was learned" summary).

Nothing here raises into the pipeline — a feedback failure must never block a run.
"""

from __future__ import annotations

import fcntl
import json
import sys
from contextlib import contextmanager
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from . import config

# The workbook column headers the user interacts with (see build_workbook).
CORRECTION_HEADER = "✎ Correct category?"
ID_HEADER = "id (don't edit)"

# How many few-shot correction examples to inject into the classifier prompt.
MAX_FEWSHOT = 12


def _now() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")


@contextmanager
def _lock(path: Path):
    path.parent.mkdir(parents=True, exist_ok=True)
    lock = path.with_name(path.name + ".lock")
    with lock.open("w") as fh:
        fcntl.flock(fh, fcntl.LOCK_EX)
        try:
            yield
        finally:
            fcntl.flock(fh, fcntl.LOCK_UN)


def _read_jsonl(path: Path) -> list[dict]:
    out: list[dict] = []
    if not path.exists():
        return out
    for line in path.read_text().split("\n"):
        if line.strip():
            try:
                out.append(json.loads(line))
            except json.JSONDecodeError:
                pass
    return out


def _atomic_write(path: Path, text: str) -> None:
    tmp = path.with_suffix(path.suffix + ".tmp")
    tmp.write_text(text)
    tmp.rename(path)


def _signature(record: dict) -> str:
    """A short, human-readable content signature for a correction (what the post
    is about), so few-shot examples read naturally and duplicates collapse."""
    title = (record.get("title") or "").strip()
    summary = (record.get("summary") or "").strip()
    sig = title or summary
    return sig[:120]


# --- reading the user's overrides -------------------------------------------

def read_workbook_overrides(workbook_path: Path | None = None) -> dict[str, str]:
    """Scan the existing workbook for filled-in 'Correct category?' cells.
    Returns {capture_id: resolved_category_key}. Empty if no workbook yet."""
    workbook_path = workbook_path or config.WORKBOOK_FILE
    if not workbook_path.exists():
        return {}
    try:
        from openpyxl import load_workbook
        wb = load_workbook(workbook_path, read_only=True, data_only=True)
    except Exception as exc:
        print(f"WARN: could not read workbook overrides: {exc}", file=sys.stderr)
        return {}
    overrides: dict[str, str] = {}
    for ws in wb.worksheets:
        header = [str(c.value) if c.value is not None else "" for c in next(ws.iter_rows(max_row=1), [])]
        if CORRECTION_HEADER not in header or ID_HEADER not in header:
            continue
        ci_corr = header.index(CORRECTION_HEADER)
        ci_id = header.index(ID_HEADER)
        for row in ws.iter_rows(min_row=2):
            if ci_id >= len(row) or ci_corr >= len(row):
                continue
            cid = row[ci_id].value
            corr = row[ci_corr].value
            if not cid or not corr or not str(corr).strip():
                continue
            key = config.resolve_category(str(corr))
            if key:
                overrides[str(cid)] = key
    wb.close()
    return overrides


def read_inbox_overrides(inbox_path: Path | None = None) -> dict[str, str]:
    """Read manual corrections from data/corrections.inbox.jsonl — lines like
    {"id": "ig_...", "to_category": "career"}. Consumed (cleared) after read."""
    inbox_path = inbox_path or config.CORRECTIONS_INBOX
    if not inbox_path.exists():
        return {}
    overrides: dict[str, str] = {}
    for row in _read_jsonl(inbox_path):
        cid, to = row.get("id"), row.get("to_category") or row.get("category")
        key = config.resolve_category(str(to)) if to else None
        if cid and key:
            overrides[str(cid)] = key
    try:
        inbox_path.unlink()  # consumed
    except OSError:
        pass
    return overrides


# --- applying corrections ----------------------------------------------------

def ingest_corrections(records: list[dict] | None = None) -> int:
    """Harvest overrides, log real corrections, pin the records. Returns count.

    MUST run before the workbook is regenerated. `records` is captures.jsonl
    already loaded; if None it is read from disk. Returns the number of NEW
    corrections applied this run."""
    from . import route_store  # local import avoids any import cycle

    if records is None:
        records = route_store._read_records(config.CAPTURES_FILE)
    by_id = {r.get("id"): r for r in records}
    if not by_id:
        return 0

    overrides = {**read_inbox_overrides(), **read_workbook_overrides()}
    if not overrides:
        return 0

    new_corrections: list[dict] = []
    updates: dict[str, dict] = {}
    for cid, to_cat in overrides.items():
        rec = by_id.get(cid)
        if not rec:
            continue
        from_cat = rec.get("category")
        if from_cat == to_cat:
            continue  # already there (e.g. user re-typed the same thing)
        new_corrections.append({
            "id": cid, "signature": _signature(rec),
            "from_category": from_cat, "to_category": to_cat, "ts": _now(),
        })
        updates[cid] = {"category": to_cat, "pinned": True,
                        "corrected_at": _now(), "corrected_from": from_cat}

    if not new_corrections:
        return 0

    _append_corrections(new_corrections)
    route_store.apply_updates(updates)  # durable, stale-write-proof, pins category
    for cid, u in updates.items():
        by_id[cid].update(u)  # keep the in-memory copy consistent
    print(f"feedback: applied {len(new_corrections)} correction(s)", file=sys.stderr)
    return len(new_corrections)


def _append_corrections(new: list[dict], path: Path | None = None) -> None:
    path = path or config.CORRECTIONS_FILE
    with _lock(path):
        existing = _read_jsonl(path)
        existing.extend(new)
        _atomic_write(path, "\n".join(json.dumps(c, ensure_ascii=True)
                                      for c in existing) + "\n")


# --- feeding corrections back into the classifier ---------------------------

def few_shot_block(path: Path | None = None) -> str:
    """Compact few-shot examples from accumulated corrections, injected into the
    classifier system prompt. Deduped, capped, short — token-conscious."""
    path = path or config.CORRECTIONS_FILE
    corrections = _read_jsonl(path)
    if not corrections:
        return ""
    seen: set[tuple] = set()
    lines: list[str] = []
    # Most recent first — recent taste wins if it flip-flopped.
    for c in reversed(corrections):
        to_cat = c.get("to_category")
        sig = (c.get("signature") or "").strip()
        if not to_cat or not sig:
            continue
        key = (sig.lower()[:60], to_cat)
        if key in seen:
            continue
        seen.add(key)
        to_label = config.pretty_label(to_cat)
        from_label = config.pretty_label(c.get("from_category") or "?")
        lines.append(f'- A post about "{sig[:80]}" should be {to_label}, not {from_label}.')
        if len(lines) >= MAX_FEWSHOT:
            break
    if not lines:
        return ""
    return ("\n\nPast corrections from the user — learn from these; classify "
            "similar posts the same way:\n" + "\n".join(lines))


# --- telemetry ---------------------------------------------------------------

def compute_metrics(records: list[dict] | None = None,
                    corrections_path: Path | None = None,
                    out_path: Path | None = None) -> dict[str, Any]:
    """Recompute data/metrics.json from captures + corrections."""
    from . import route_store
    corrections_path = corrections_path or config.CORRECTIONS_FILE
    out_path = out_path or config.METRICS_FILE
    if records is None:
        records = route_store._read_records(config.CAPTURES_FILE)
    corrections = _read_jsonl(corrections_path)

    total_items = len(records)
    total_corr = len(corrections)

    by_transition: dict[str, int] = {}
    from_counts: dict[str, int] = {}
    for c in corrections:
        f = config.pretty_label(c.get("from_category") or "?")
        t = config.pretty_label(c.get("to_category") or "?")
        by_transition[f"{f} → {t}"] = by_transition.get(f"{f} → {t}", 0) + 1
        from_counts[f] = from_counts.get(f, 0) + 1

    # Monthly correction rate (should trend DOWN as the classifier learns).
    items_by_month: dict[str, int] = {}
    for r in records:
        m = (r.get("ts") or "")[:7]
        if m:
            items_by_month[m] = items_by_month.get(m, 0) + 1
    corr_by_month: dict[str, int] = {}
    for c in corrections:
        m = (c.get("ts") or "")[:7]
        if m:
            corr_by_month[m] = corr_by_month.get(m, 0) + 1
    rate_over_time = []
    for m in sorted(set(items_by_month) | set(corr_by_month)):
        items = items_by_month.get(m, 0)
        corr = corr_by_month.get(m, 0)
        rate_over_time.append({"month": m, "items": items, "corrections": corr,
                               "rate": round(corr / items, 3) if items else None})

    most_corrected = sorted(from_counts.items(), key=lambda kv: -kv[1])
    learned = _learned_summary(total_corr, by_transition, rate_over_time)

    metrics = {
        "generated_at": _now(),
        "total_items_classified": total_items,
        "total_corrections": total_corr,
        "correction_rate": round(total_corr / total_items, 3) if total_items else 0.0,
        "by_transition": dict(sorted(by_transition.items(), key=lambda kv: -kv[1])),
        "most_corrected_from": dict(most_corrected),
        "rate_over_time": rate_over_time,
        "learned": learned,
    }
    try:
        out_path.parent.mkdir(parents=True, exist_ok=True)
        _atomic_write(out_path, json.dumps(metrics, ensure_ascii=True, indent=2))
    except Exception as exc:
        print(f"WARN: could not write metrics: {exc}", file=sys.stderr)
    return metrics


def _learned_summary(total_corr, by_transition, rate_over_time) -> str:
    if total_corr == 0:
        return ("No corrections yet. Fix a category in the workbook's "
                f"'{CORRECTION_HEADER}' column and the classifier will learn from it.")
    top = next(iter(by_transition.items()), None)
    parts = [f"{total_corr} correction(s) captured and being used as classifier hints."]
    if top:
        parts.append(f"Most common fix: {top[0]} ({top[1]}×).")
    rated = [r for r in rate_over_time if r["rate"] is not None]
    if len(rated) >= 2 and rated[-1]["rate"] < rated[0]["rate"]:
        parts.append(f"Correction rate is trending down "
                     f"({rated[0]['rate']:.0%} → {rated[-1]['rate']:.0%}) — it's learning.")
    return " ".join(parts)
