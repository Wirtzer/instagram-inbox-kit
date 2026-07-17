"""Build the Excel workbook from captures.jsonl — one sheet per collection or
content category, each with a purpose-built column layout.

Tab rule: a saved-to collection name wins (e.g. a "Restaurants" or game
collection); otherwise the content category (Recipes, Movies, …). Books and
Restaurants also get deduped master tabs. Deterministic full rebuild; safe to
run any time.

    python -m ig_inbox.build_workbook
"""

from __future__ import annotations

import json
import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Callable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from . import config
from .feedback import CORRECTION_HEADER, ID_HEADER

HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(bold=True, color="FFFFFF")
CORRECT_FILL = PatternFill("solid", fgColor="B45309")  # amber: "edit me"
TITLE_FONT = Font(bold=True, size=13)
WRAP = Alignment(wrap_text=True, vertical="top")

# sheet names come from config.pretty_label (any category, incl. coined ones)


def _load(captures: Path) -> list[dict]:
    if not captures.exists():
        return []
    recs = []
    for line in captures.read_text().split("\n"):
        if line.strip():
            try:
                recs.append(json.loads(line))
            except json.JSONDecodeError:
                pass
    return recs


def _date(r: dict) -> str:
    return (r.get("ts") or "")[:10]


def _src(r: dict) -> str:
    return r.get("permalink") or r.get("url") or ""


def _join(v: Any) -> str:
    if isinstance(v, list):
        return "\n".join(f"• {x}" for x in v)
    return str(v or "")


# --- per-category column schemas: (header, value) ---------------------------

def _common_cols(r: dict) -> dict:
    return {"Date": _date(r), "From": f"@{r.get('author') or '—'}", "Link": _src(r)}


def cols_recipe(r: dict) -> list[tuple[str, str]]:
    d = r.get("details") or {}
    return [("Dish", r.get("title") or ""), ("Cuisine", d.get("cuisine", "")),
            ("Ingredients to buy", _join(d.get("ingredients"))),
            ("How to make it", _join(d.get("steps"))),
            ("Summary", r.get("summary", ""))]


def cols_restaurant(r: dict) -> list[tuple[str, str]]:
    d = r.get("details") or {}
    e = r.get("enrichment") or {}
    return [("Restaurant", e.get("full_name") or r.get("title") or d.get("name") or ""),
            ("Cuisine / type", e.get("cuisine") or d.get("cuisine", "")),
            ("Address", e.get("address", "")), ("City", e.get("city", "")),
            ("Hours", e.get("hours", "")), ("Rating", e.get("rating", "")),
            ("Price", e.get("price", "")),
            ("What the post says", d.get("notes") or r.get("summary", "")),
            ("Website", e.get("website", ""))]


def cols_content(r: dict) -> list[tuple[str, str]]:
    """Game / how-to / AI posts: the ACTUAL info, not a shallow summary.
    key_points = extracted facts; links = repos/tools; assessment = the verdict."""
    kp = r.get("key_points") or []
    body = "\n".join(f"• {p}" for p in kp) if kp else (r.get("summary") or "")
    cols = [("What it is", r.get("title") or ""),
            ("Key info from the post", body),
            ("Links", "\n".join(r.get("links") or []))]
    if r.get("category") == "ai_idea":
        cols.append(("Assessment (fit for you?)", r.get("assessment", "")))
    cols.append(("Summary", r.get("summary", "")))
    return cols


def cols_movie(r: dict) -> list[tuple[str, str]]:
    d = r.get("details") or {}
    return [("Title", r.get("title") or d.get("title", "")), ("Year", d.get("year", "")),
            ("Genre", d.get("genre", "")), ("Where to watch", d.get("where_to_watch", "")),
            ("What the post says", r.get("summary", ""))]


def cols_random_food(r: dict) -> list[tuple[str, str]]:
    d = r.get("details") or {}
    return [("Topic", d.get("topic") or r.get("title") or ""),
            ("The picks", _join(d.get("picks"))), ("Summary", r.get("summary", ""))]


def cols_book(r: dict) -> list[tuple[str, str]]:
    books = (r.get("entities") or {}).get("books") or []
    names = "\n".join(f"• {b['title']}" + (f" — {b['author']}" if b.get("author") else "")
                      for b in books)
    return [("Books", names or r.get("title", "")), ("Why", r.get("summary", ""))]


def cols_generic(r: dict) -> list[tuple[str, str]]:
    d = r.get("details") or {}
    extra = "; ".join(f"{k}: {_join(v)}" for k, v in d.items() if v)
    return [("What it is", r.get("title") or ""), ("Summary", r.get("summary", "")),
            ("Details", extra)]


SCHEMA: dict[str, Callable[[dict], list[tuple[str, str]]]] = {
    "recipe": cols_recipe, "restaurant": cols_restaurant, "movie": cols_movie,
    "random_food": cols_random_food, "book": cols_book,
    "game": cols_content, "ai_idea": cols_content, "research": cols_content,
    "product": cols_content, "self_improvement": cols_content,
    "relationships": cols_content, "career": cols_content, "news": cols_content,
}


def _safe_sheet_name(name: str, taken: set[str]) -> str:
    clean = re.sub(r"[\[\]:*?/\\]", "", name).strip()[:31] or "Sheet"
    base, i = clean, 2
    while clean.lower() in taken:
        suffix = f" {i}"
        clean = base[:31 - len(suffix)] + suffix
        i += 1
    taken.add(clean.lower())
    return clean


def _write_sheet(wb: Workbook, sheet_name: str, records: list[dict],
                 col_fn: Callable[[dict], list[tuple[str, str]]]) -> None:
    ws = wb.create_sheet(sheet_name)
    records = sorted(records, key=lambda r: r.get("ts", ""), reverse=True)
    # The two trailing feedback columns (CORRECTION_HEADER + ID_HEADER) are how a
    # user re-categorizes: type a category into the amber column, save. The next
    # run reads it back BEFORE regenerating and learns from it.
    headers = ([h for h, _ in col_fn(records[0])]
               + ["Date", "From", "Link", CORRECTION_HEADER, ID_HEADER])
    ws.append(headers)
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci)
        c.fill = CORRECT_FILL if h == CORRECTION_HEADER else HEADER_FILL
        c.font = HEADER_FONT
    for r in records:
        vals = [v for _, v in col_fn(r)]
        common = _common_cols(r)
        ws.append(vals + [common["Date"], common["From"], common["Link"], "", r.get("id", "")])
    widths = {"Ingredients to buy": 40, "Key info from the post": 55,
              "Assessment (fit for you?)": 50, "Links": 30, "How to make it": 50,
              "The picks": 40, "What the post says": 45, "Summary": 45,
              "Address": 30, "Books": 40, "Details": 40, "Link": 22, "Website": 26,
              CORRECTION_HEADER: 20, ID_HEADER: 22}
    for ci, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(ci)].width = widths.get(h, 18)
        for ri in range(2, ws.max_row + 1):
            ws.cell(row=ri, column=ci).alignment = WRAP
    ws.freeze_panes = "A2"


def _write_metrics(wb: Workbook, metrics: dict, taken: set[str]) -> None:
    """Telemetry tab: totals, transitions, correction rate over time."""
    ws = wb.create_sheet(_safe_sheet_name("Metrics", taken))
    ws["A1"] = "Learning metrics"
    ws["A1"].font = TITLE_FONT
    rows = [
        ("Items classified", metrics.get("total_items_classified", 0)),
        ("Total corrections", metrics.get("total_corrections", 0)),
        ("Correction rate", metrics.get("correction_rate", 0.0)),
    ]
    r = 3
    for label, val in rows:
        ws.cell(row=r, column=1, value=label).font = Font(bold=True)
        ws.cell(row=r, column=2, value=val)
        r += 1
    ws.cell(row=r + 1, column=1, value="What was learned").font = Font(bold=True)
    ws.cell(row=r + 1, column=2, value=metrics.get("learned", "")).alignment = WRAP

    # Transitions table
    tr = r + 4
    ws.cell(row=tr, column=1, value="Correction (from → to)").fill = HEADER_FILL
    ws.cell(row=tr, column=1).font = HEADER_FONT
    ws.cell(row=tr, column=2, value="Count").fill = HEADER_FILL
    ws.cell(row=tr, column=2).font = HEADER_FONT
    tr += 1
    for k, v in (metrics.get("by_transition") or {}).items():
        ws.cell(row=tr, column=1, value=k)
        ws.cell(row=tr, column=2, value=v)
        tr += 1

    # Correction rate over time (should trend down as it learns)
    ot = tr + 2
    for ci, h in enumerate(("Month", "Items", "Corrections", "Rate"), 1):
        c = ws.cell(row=ot, column=ci, value=h)
        c.fill, c.font = HEADER_FILL, HEADER_FONT
    ot += 1
    for row in metrics.get("rate_over_time") or []:
        ws.cell(row=ot, column=1, value=row.get("month"))
        ws.cell(row=ot, column=2, value=row.get("items"))
        ws.cell(row=ot, column=3, value=row.get("corrections"))
        ws.cell(row=ot, column=4, value=row.get("rate"))
        ot += 1
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 50
    ws.cell(row=r + 1, column=2).alignment = WRAP


def _write_books_master(wb: Workbook, taken: set[str], books_master: Path) -> int:
    """Deduped master of EVERY book across all posts."""
    if not books_master.exists():
        return 0
    try:
        master = json.loads(books_master.read_text())
    except json.JSONDecodeError:
        return 0
    books = sorted(master.values(), key=lambda b: (b.get("title") or "").lower())
    ws = wb.create_sheet(_safe_sheet_name("Books", taken))
    headers = ["Title", "Author", "Genre", "Rating", "About",
               "Recommended by", "Why recommended", "# posts", "Sources"]
    ws.append(headers)
    for ci in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=ci); c.fill, c.font = HEADER_FILL, HEADER_FONT
    for b in books:
        srcs = b.get("sources") or []
        recs = b.get("recommenders") or []
        handles = "\n".join(dict.fromkeys(
            f"@{r['handle']}" for r in recs if r.get("handle")))
        whys = "\n".join(dict.fromkeys(
            f"“{r['why']}”" + (f" — @{r['handle']}" if r.get("handle") else "")
            for r in recs if r.get("why")))
        ws.append([b.get("title", ""), b.get("author", ""), b.get("genre", ""),
                   b.get("rating", ""), b.get("about", ""), handles, whys,
                   len(srcs), "\n".join(srcs)])
    for ci, w in enumerate([32, 22, 18, 14, 46, 20, 40, 8, 24], 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
        for ri in range(2, ws.max_row + 1):
            ws.cell(row=ri, column=ci).alignment = WRAP
    ws.freeze_panes = "A2"
    return len(books)


def _write_restaurants_master(wb: Workbook, recs: list[dict], taken: set[str]) -> int:
    """Deduped master of EVERY restaurant mentioned (one row per unique place)."""
    by_name: dict[str, dict] = {}
    for r in recs:
        if r.get("category") != "restaurant":
            continue
        e = r.get("enrichment") or {}
        d = r.get("details") or {}
        name = e.get("full_name") or r.get("title") or d.get("name") or ""
        key = re.sub(r"[^a-z0-9]+", "", name.lower())
        if not key:
            continue
        src = r.get("permalink") or r.get("url") or ""
        if key not in by_name:
            by_name[key] = {"name": name, "cuisine": e.get("cuisine") or d.get("cuisine", ""),
                            "address": e.get("address", ""), "city": e.get("city", ""),
                            "hours": e.get("hours", ""), "rating": e.get("rating", ""),
                            "price": e.get("price", ""), "website": e.get("website", ""),
                            "notes": d.get("notes") or r.get("summary", ""), "sources": []}
        row = by_name[key]
        for f in ("cuisine", "address", "city", "hours", "rating", "price", "website"):
            if not row[f] and (e.get(f) or d.get(f)):
                row[f] = e.get(f) or d.get(f)
        if src and src not in row["sources"]:
            row["sources"].append(src)
    ws = wb.create_sheet(_safe_sheet_name("Restaurants", taken))
    headers = ["Restaurant", "Cuisine / type", "Address", "City", "Hours", "Rating",
               "Price", "Notes", "Website", "# posts", "Sources"]
    ws.append(headers)
    for ci in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=ci); c.fill, c.font = HEADER_FILL, HEADER_FONT
    for row in sorted(by_name.values(), key=lambda x: x["name"].lower()):
        s = row["sources"]
        ws.append([row["name"], row["cuisine"], row["address"], row["city"], row["hours"],
                   row["rating"], row["price"], row["notes"], row["website"], len(s), "\n".join(s)])
    for ci, w in enumerate([26, 20, 30, 14, 20, 8, 8, 40, 24, 8, 24], 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
        for ri in range(2, ws.max_row + 1):
            ws.cell(row=ri, column=ci).alignment = WRAP
    ws.freeze_panes = "A2"
    return len(by_name)


def _game_tab(rec: dict, known_games: dict[str, str]) -> str | None:
    """For a game post, the canonical game name to use as its tab (so variants
    all land in one tab instead of scattering)."""
    hay = f"{rec.get('title','')} {(rec.get('details') or {}).get('notes','')} " \
          f"{rec.get('collection','')} {rec.get('summary','')}".lower()
    for needle, canon in known_games.items():
        if needle in hay:
            return canon
    return None


def build(captures_path: Path | None = None, out_path: Path | None = None,
          books_master: Path | None = None, known_games: dict | None = None,
          metrics: dict | None = None) -> int:
    captures_path = captures_path or config.CAPTURES_FILE
    out_path = out_path or config.WORKBOOK_FILE
    books_master = books_master or config.BOOKS_MASTER_FILE
    if known_games is None:
        known_games = config.load_config().get("known_games") or {}

    recs = _load(captures_path)
    if metrics is None:
        try:
            from . import feedback
            metrics = feedback.compute_metrics(records=recs)
        except Exception:
            metrics = {}
    wb = Workbook()
    wb.remove(wb.active)
    taken = {"overview"}

    ov = wb.create_sheet("Overview")
    ov["A1"] = "Instagram Inbox · Captures"
    ov["A1"].font = TITLE_FONT
    ov["A2"] = (f"Auto-updated {datetime.now(timezone.utc):%Y-%m-%d %H:%M UTC} · "
                f"{len(recs)} posts. Books & Restaurants are deduped master lists; "
                "other tabs are per-post with the extracted details.")

    if metrics:
        _write_metrics(wb, metrics, taken)
    n_books = _write_books_master(wb, taken, books_master)
    n_rest = _write_restaurants_master(wb, recs, taken)

    groups: dict[str, list[dict]] = {}
    order: list[str] = []
    for r in recs:
        if r.get("category") in ("book", "restaurant") and not (r.get("collection") or "").strip():
            continue  # already in a master
        if r.get("category") == "game":
            name = _game_tab(r, known_games) or "Games"
        else:
            coll = (r.get("collection") or "").strip()
            name = coll or config.pretty_label(r.get("category", "other"))
        if name in ("Books", "Restaurants"):
            continue  # collection literally named Books/Restaurants → the master
        if name not in groups:
            groups[name] = []
            order.append(name)
        groups[name].append(r)

    ov.append([]); ov.append(["Tab", "Items"])
    for c in ("A4", "B4"):
        ov[c].font = HEADER_FONT; ov[c].fill = HEADER_FILL
    ov.append(["Books (master)", n_books])
    ov.append(["Restaurants (master)", n_rest])
    for name in sorted(order, key=lambda n: -len(groups[n])):
        ov.append([name, len(groups[name])])
    ov.column_dimensions["A"].width = 26
    ov.column_dimensions["B"].width = 10

    is_collection = {n for n in order if any((r.get("collection") or "").strip() == n
                                             for r in groups[n])}
    for name in sorted(order, key=lambda n: (n not in is_collection, -len(groups[n]))):
        cats = [r.get("category") for r in groups[name]]
        dom = max(set(cats), key=cats.count) if cats else "other"
        col_fn = SCHEMA.get(dom, cols_generic)
        _write_sheet(wb, _safe_sheet_name(name, taken), groups[name], col_fn)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return len(recs)


if __name__ == "__main__":
    print(f"workbook built from {build()} captures → {config.WORKBOOK_FILE}")
